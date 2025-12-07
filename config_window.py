import sys
import os
import json
import shutil
from datetime import datetime
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGroupBox, QLabel, QLineEdit, QPushButton, QComboBox, QSpinBox,
    QTextEdit, QFileDialog, QMessageBox, QTableWidget, QTableWidgetItem,
    QRadioButton, QButtonGroup
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QIcon, QFont
import openpyxl


class ConfigWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.workbook = None
        self.sheet_names = []
        self.max_columns = 26  # Default A-Z
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("Barcode Scanner Configuration")
        self.setGeometry(100, 100, 900, 800)
        
        # Main widget and layout
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)
        
        # File selection section
        file_group = QGroupBox("Excel File Selection")
        file_layout = QHBoxLayout()
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setText(r'EXCEL SHEET')
        file_browse_btn = QPushButton("Browse...")
        file_browse_btn.clicked.connect(self.browse_file)
        file_layout.addWidget(QLabel("File Path:"))
        file_layout.addWidget(self.file_path_edit)
        file_layout.addWidget(file_browse_btn)
        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)
        
        # Sheet selection section
        sheet_group = QGroupBox("Sheet Names")
        sheet_layout = QVBoxLayout()
        
        self.inventory_sheet_group = QButtonGroup()
        self.other_sheet_group = QButtonGroup()
        
        inv_label = QLabel("Inventory Sheet:")
        inv_label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        sheet_layout.addWidget(inv_label)
        
        self.inventory_radio_layout = QVBoxLayout()
        sheet_layout.addLayout(self.inventory_radio_layout)
        
        other_label = QLabel("Other Sheet:")
        other_label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        sheet_layout.addWidget(other_label)
        
        self.other_radio_layout = QVBoxLayout()
        sheet_layout.addLayout(self.other_radio_layout)
        
        custom_sheet_layout = QHBoxLayout()
        custom_sheet_layout.addWidget(QLabel("Add Custom Sheet Name:"))
        self.custom_inventory_sheet = QLineEdit()
        self.custom_inventory_sheet.setPlaceholderText("Custom inventory sheet name")
        custom_sheet_layout.addWidget(self.custom_inventory_sheet)
        self.custom_other_sheet = QLineEdit()
        self.custom_other_sheet.setPlaceholderText("Custom other sheet name")
        custom_sheet_layout.addWidget(self.custom_other_sheet)
        sheet_layout.addLayout(custom_sheet_layout)
        
        sheet_group.setLayout(sheet_layout)
        main_layout.addWidget(sheet_group)
        
        # Column selection section
        column_group = QGroupBox("Column Mappings")
        column_layout = QVBoxLayout()
        
        # Asset ID columns (3 dropdowns with "None" option)
        asset_id_layout = QHBoxLayout()
        asset_id_layout.addWidget(QLabel("Asset ID Columns (1-3):"))
        self.asset_id1_combo = QComboBox()
        self.asset_id2_combo = QComboBox()
        self.asset_id3_combo = QComboBox()
        asset_id_layout.addWidget(self.asset_id1_combo)
        asset_id_layout.addWidget(self.asset_id2_combo)
        asset_id_layout.addWidget(self.asset_id3_combo)
        column_layout.addLayout(asset_id_layout)
        
        # Required columns
        self.asset_name_combo = self.create_column_row(column_layout, "Asset Name Column:")
        self.asset_desc_combo = self.create_column_row(column_layout, "Asset Description Column:")
        self.status_combo = self.create_column_row(column_layout, "Status Column:")
        self.location_combo = self.create_column_row(column_layout, "Building/Location Column:")
        self.room_combo = self.create_column_row(column_layout, "Room Column:")
        self.marked_check_combo = self.create_column_row(column_layout, "Marked Check Column:")
        
        column_group.setLayout(column_layout)
        main_layout.addWidget(column_group)
        
        # Row range section
        range_group = QGroupBox("Row Range for Counting")
        range_layout = QHBoxLayout()
        range_layout.addWidget(QLabel("Start Row:"))
        self.start_row_spin = QSpinBox()
        self.start_row_spin.setMinimum(1)
        self.start_row_spin.setMaximum(1000000)
        self.start_row_spin.setValue(6)
        self.start_row_spin.valueChanged.connect(self.update_total_count)
        range_layout.addWidget(self.start_row_spin)
        
        range_layout.addWidget(QLabel("End Row:"))
        self.end_row_spin = QSpinBox()
        self.end_row_spin.setMinimum(1)
        self.end_row_spin.setMaximum(1000000)
        self.end_row_spin.setValue(357)
        self.end_row_spin.valueChanged.connect(self.update_total_count)
        range_layout.addWidget(self.end_row_spin)
        
        self.total_count_label = QLabel("Total Count: 352 items")
        self.total_count_label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        range_layout.addWidget(self.total_count_label)
        range_layout.addStretch()
        
        range_group.setLayout(range_layout)
        main_layout.addWidget(range_group)
        
        # ngrok URL section
        ngrok_group = QGroupBox("Server Configuration")
        ngrok_layout = QHBoxLayout()
        ngrok_layout.addWidget(QLabel("ngrok Server URL:"))
        self.ngrok_url_edit = QLineEdit()
        self.ngrok_url_edit.setPlaceholderText("https://example.ngrok-free.dev/api")
        self.ngrok_url_edit.setText("<PASTE_NGROK_ADDRESS_HERE>/api")
        ngrok_layout.addWidget(self.ngrok_url_edit)
        ngrok_group.setLayout(ngrok_layout)
        main_layout.addWidget(ngrok_group)
        
        # Validation status section
        validation_layout = QHBoxLayout()
        self.validation_label = QLabel("⚪ Configuration not validated")
        validation_layout.addWidget(self.validation_label)
        validation_layout.addStretch()
        main_layout.addLayout(validation_layout)
        
        # Summary section
        summary_group = QGroupBox("Configuration Summary")
        summary_layout = QVBoxLayout()
        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        self.summary_text.setMaximumHeight(150)
        summary_layout.addWidget(self.summary_text)
        summary_group.setLayout(summary_layout)
        main_layout.addWidget(summary_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        test_btn = QPushButton("Test Configuration")
        test_btn.clicked.connect(self.test_configuration)
        save_btn = QPushButton("Save Configuration")
        save_btn.clicked.connect(self.save_configuration)
        button_layout.addWidget(test_btn)
        button_layout.addWidget(save_btn)
        button_layout.addStretch()
        main_layout.addLayout(button_layout)
        
        # Initialize columns
        self.populate_column_dropdowns()
        
        # Load Excel file if it exists
        if os.path.exists(self.file_path_edit.text()):
            self.load_excel_file(self.file_path_edit.text())
        
        # Connect signals for real-time validation
        self.file_path_edit.textChanged.connect(self.on_file_path_changed)
        self.asset_id1_combo.currentTextChanged.connect(self.validate_and_update_summary)
        self.asset_id2_combo.currentTextChanged.connect(self.validate_and_update_summary)
        self.asset_id3_combo.currentTextChanged.connect(self.validate_and_update_summary)
        self.asset_name_combo.currentTextChanged.connect(self.validate_and_update_summary)
        self.asset_desc_combo.currentTextChanged.connect(self.validate_and_update_summary)
        self.status_combo.currentTextChanged.connect(self.validate_and_update_summary)
        self.location_combo.currentTextChanged.connect(self.validate_and_update_summary)
        self.room_combo.currentTextChanged.connect(self.validate_and_update_summary)
        self.marked_check_combo.currentTextChanged.connect(self.validate_and_update_summary)
        
        # Set default selections
        self.set_default_selections()
        
        # Load existing config if it exists
        self.load_existing_config()
        
    def create_column_row(self, parent_layout, label_text):
        layout = QHBoxLayout()
        layout.addWidget(QLabel(label_text))
        combo = QComboBox()
        layout.addWidget(combo)
        layout.addStretch()
        parent_layout.addLayout(layout)
        return combo
        
    def populate_column_dropdowns(self):
        """Populate column dropdowns with A-ZZ"""
        columns = self.generate_column_list()
        
        # Asset ID dropdowns include "None" option
        for combo in [self.asset_id1_combo, self.asset_id2_combo, self.asset_id3_combo]:
            combo.clear()
            combo.addItem("None")
            combo.addItems(columns)
            
        # Required dropdowns (no "None" option)
        for combo in [self.asset_name_combo, self.asset_desc_combo, self.status_combo,
                      self.location_combo, self.room_combo, self.marked_check_combo]:
            combo.clear()
            combo.addItems(columns)
            
    def generate_column_list(self):
        """Generate column letters A-ZZ"""
        columns = []
        # A-Z (0-25)
        for i in range(26):
            columns.append(chr(65 + i))
        # AA-ZZ (26-701)
        for i in range(26):
            for j in range(26):
                columns.append(chr(65 + i) + chr(65 + j))
        return columns
        
    def letter_to_index(self, col_letter):
        """Convert column letter to zero-based index"""
        if not col_letter or col_letter == "None":
            return None
        col_letter = col_letter.strip().upper()
        if not col_letter:
            return None
        if len(col_letter) == 1:
            return ord(col_letter) - 65
        else:
            return (ord(col_letter[0]) - 65 + 1) * 26 + ord(col_letter[1]) - 65
            
    def index_to_letter(self, index):
        """Convert zero-based index to column letter"""
        if index is None:
            return "None"
        if index < 26:
            return chr(65 + index)
        else:
            first = (index - 26) // 26
            second = (index - 26) % 26
            return chr(65 + first) + chr(65 + second)
    
    def load_existing_config(self):
        """Load existing config.json if it exists"""
        config_path = os.path.join(os.path.dirname(__file__), "config.json")
        if not os.path.exists(config_path):
            return
            
        try:
            with open(config_path, 'r') as f:
                config = json.load(f)
                
            # Load Excel file path
            if "excel" in config and "filePath" in config["excel"]:
                self.file_path_edit.setText(config["excel"]["filePath"])
                
            # Load column mappings
            if "excel" in config and "columns" in config["excel"]:
                cols = config["excel"]["columns"]
                
                # Asset ID columns
                asset_ids = cols.get("assetIdSearch", [])
                if len(asset_ids) > 0 and asset_ids[0] is not None:
                    self.asset_id1_combo.setCurrentText(self.index_to_letter(asset_ids[0]))
                if len(asset_ids) > 1 and asset_ids[1] is not None:
                    self.asset_id2_combo.setCurrentText(self.index_to_letter(asset_ids[1]))
                if len(asset_ids) > 2 and asset_ids[2] is not None:
                    self.asset_id3_combo.setCurrentText(self.index_to_letter(asset_ids[2]))
                    
                # Other columns
                if cols.get("assetName") is not None:
                    self.asset_name_combo.setCurrentText(self.index_to_letter(cols["assetName"]))
                if cols.get("assetDescription") is not None:
                    self.asset_desc_combo.setCurrentText(self.index_to_letter(cols["assetDescription"]))
                if cols.get("status") is not None:
                    self.status_combo.setCurrentText(self.index_to_letter(cols["status"]))
                if cols.get("location") is not None:
                    self.location_combo.setCurrentText(self.index_to_letter(cols["location"]))
                if cols.get("room") is not None:
                    self.room_combo.setCurrentText(self.index_to_letter(cols["room"]))
                if cols.get("markedCheck") is not None:
                    self.marked_check_combo.setCurrentText(self.index_to_letter(cols["markedCheck"]))
                    
            # Load row range
            if "excel" in config and "counting" in config["excel"]:
                counting = config["excel"]["counting"]
                if "startRow" in counting:
                    self.start_row_spin.setValue(counting["startRow"])
                if "endRow" in counting:
                    self.end_row_spin.setValue(counting["endRow"])
                    
            # Load ngrok URL
            if "server" in config and "ngrokUrl" in config["server"]:
                self.ngrok_url_edit.setText(config["server"]["ngrokUrl"])
                
            self.validation_label.setText("✅ Loaded existing configuration")
            
        except Exception as e:
            self.validation_label.setText(f"⚠️ Error loading config: {str(e)}")
            
    def browse_file(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if file_name:
            self.file_path_edit.setText(file_name)
            
    def on_file_path_changed(self):
        file_path = self.file_path_edit.text()
        if os.path.exists(file_path):
            self.load_excel_file(file_path)
        self.validate_and_update_summary()
            
    def load_excel_file(self, file_path):
        """Load Excel file and detect sheets"""
        try:
            self.workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            self.sheet_names = self.workbook.sheetnames
            
            # Populate sheet radio buttons
            self.populate_sheet_radios()
            
            # Get max columns from first sheet
            if self.sheet_names:
                ws = self.workbook[self.sheet_names[0]]
                self.max_columns = ws.max_column
                self.populate_column_dropdowns()
                
            self.validation_label.setText("✅ Excel file loaded successfully")
        except Exception as e:
            self.validation_label.setText(f"❌ Error loading Excel file: {str(e)}")
            
    def populate_sheet_radios(self):
        """Populate radio buttons for sheet selection"""
        # Clear existing radios
        for i in reversed(range(self.inventory_radio_layout.count())):
            self.inventory_radio_layout.itemAt(i).widget().deleteLater()
        for i in reversed(range(self.other_radio_layout.count())):
            self.other_radio_layout.itemAt(i).widget().deleteLater()
            
        # Create radio buttons for inventory sheet
        for sheet_name in self.sheet_names:
            radio = QRadioButton(sheet_name)
            self.inventory_sheet_group.addButton(radio)
            self.inventory_radio_layout.addWidget(radio)
            if sheet_name == "Sheet1":
                radio.setChecked(True)
                
        # Create radio buttons for other sheet
        for sheet_name in self.sheet_names:
            radio = QRadioButton(sheet_name)
            self.other_sheet_group.addButton(radio)
            self.other_radio_layout.addWidget(radio)
            if sheet_name == "Other":
                radio.setChecked(True)
                
    def set_default_selections(self):
        """Set default column selections based on current hardcoded values"""
        # Asset ID: C, D, E (indices 2, 3, 4)
        self.asset_id1_combo.setCurrentText("C")
        self.asset_id2_combo.setCurrentText("D")
        self.asset_id3_combo.setCurrentText("E")
        
        # Asset Name: F (index 5)
        self.asset_name_combo.setCurrentText("F")
        
        # Asset Description: G (index 6)
        self.asset_desc_combo.setCurrentText("G")
        
        # Status: P (index 15)
        self.status_combo.setCurrentText("P")
        
        # Location: Q (index 16) - actually T based on server.js
        self.location_combo.setCurrentText("Q")
        
        # Room: R (index 17) - actually U based on server.js
        self.room_combo.setCurrentText("R")
        
        # Marked Check: S (index 18)
        self.marked_check_combo.setCurrentText("S")
        
    def update_total_count(self):
        """Update total count label based on row range"""
        start = self.start_row_spin.value()
        end = self.end_row_spin.value()
        total = max(0, end - start + 1)
        self.total_count_label.setText(f"Total Count: {total} items")
        self.validate_and_update_summary()
        
    def validate_and_update_summary(self):
        """Validate configuration and update summary in real-time"""
        errors = []
        warnings = []
        
        # Check if file exists
        if not os.path.exists(self.file_path_edit.text()):
            errors.append("Excel file does not exist")
            
        # Get selected columns
        asset_id_cols = []
        for combo in [self.asset_id1_combo, self.asset_id2_combo, self.asset_id3_combo]:
            col = combo.currentText()
            if col != "None":
                idx = self.letter_to_index(col)
                if idx is not None and idx < 702:  # A-ZZ range
                    asset_id_cols.append((col, idx))
                else:
                    errors.append(f"Invalid Asset ID column: {col}")
                    
        # Check required columns
        required_combos = {
            "Asset Name": self.asset_name_combo,
            "Asset Description": self.asset_desc_combo,
            "Status": self.status_combo,
            "Location": self.location_combo,
            "Room": self.room_combo,
            "Marked Check": self.marked_check_combo
        }
        
        required_cols = {}
        for name, combo in required_combos.items():
            col = combo.currentText()
            idx = self.letter_to_index(col)
            if idx is None or idx >= 702:
                errors.append(f"Invalid {name} column: {col}")
            else:
                required_cols[name] = (col, idx)
                
        # Check for overlaps between Asset ID and other columns
        asset_id_indices = [idx for _, idx in asset_id_cols]
        for name, (col, idx) in required_cols.items():
            if idx in asset_id_indices:
                warnings.append(f"Warning: {name} column {col} overlaps with Asset ID columns")
                
        # Update validation label
        if errors:
            self.validation_label.setText("❌ " + "; ".join(errors))
        elif warnings:
            self.validation_label.setText("⚠️ " + "; ".join(warnings))
        else:
            self.validation_label.setText("✅ Configuration valid")
            
        # Update summary text
        summary = []
        asset_id_str = ", ".join([col for col, _ in asset_id_cols]) if asset_id_cols else "None"
        asset_id_idx_str = ", ".join([str(idx) for _, idx in asset_id_cols]) if asset_id_cols else "None"
        summary.append(f"Asset ID Columns: {asset_id_str} (indices: {asset_id_idx_str})")
        
        for name, (col, idx) in required_cols.items():
            summary.append(f"{name}: {col} (index: {idx})")
            
        summary.append(f"Row Range: {self.start_row_spin.value()}-{self.end_row_spin.value()} (Total: {max(0, self.end_row_spin.value() - self.start_row_spin.value() + 1)} items)")
        summary.append(f"ngrok URL: {self.ngrok_url_edit.text()}")
        
        self.summary_text.setPlainText("\n".join(summary))
        
    def test_configuration(self):
        """Test configuration by reading sample data from Excel"""
        file_path = self.file_path_edit.text()
        
        if not os.path.exists(file_path):
            QMessageBox.critical(self, "Error", "Excel file does not exist!")
            return
            
        try:
            # Get selected inventory sheet
            inventory_sheet = None
            for button in self.inventory_sheet_group.buttons():
                if button.isChecked():
                    inventory_sheet = button.text()
                    break
                    
            if not inventory_sheet and self.custom_inventory_sheet.text():
                inventory_sheet = self.custom_inventory_sheet.text()
                
            if not inventory_sheet:
                QMessageBox.critical(self, "Error", "No inventory sheet selected!")
                return
                
            # Open workbook and read sample data
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            if inventory_sheet not in wb.sheetnames:
                QMessageBox.warning(self, "Warning", f"Sheet '{inventory_sheet}' not found in Excel file. It will be created when needed.")
                return
                
            ws = wb[inventory_sheet]
            
            # Get selected columns
            columns_to_read = []
            column_labels = []
            
            for combo in [self.asset_id1_combo, self.asset_id2_combo, self.asset_id3_combo]:
                col = combo.currentText()
                if col != "None":
                    idx = self.letter_to_index(col)
                    columns_to_read.append(idx + 1)  # openpyxl uses 1-based indexing
                    column_labels.append(f"Asset ID ({col})")
                    
            for combo, label in [(self.asset_name_combo, "Name"), (self.asset_desc_combo, "Desc"),
                                  (self.status_combo, "Status"), (self.location_combo, "Location"),
                                  (self.room_combo, "Room"), (self.marked_check_combo, "Marked")]:
                col = combo.currentText()
                idx = self.letter_to_index(col)
                columns_to_read.append(idx + 1)
                column_labels.append(f"{label} ({col})")
                
            # Create preview dialog
            preview_dialog = QMessageBox(self)
            preview_dialog.setWindowTitle("Configuration Test - Sample Data")
            
            # Create table widget
            table = QTableWidget()
            table.setRowCount(5)
            table.setColumnCount(len(columns_to_read))
            table.setHorizontalHeaderLabels(column_labels)
            
            # Read first 5 rows
            for row in range(1, 6):
                for col_idx, excel_col in enumerate(columns_to_read):
                    cell_value = ws.cell(row=row, column=excel_col).value
                    table.setItem(row - 1, col_idx, QTableWidgetItem(str(cell_value) if cell_value is not None else ""))
                    
            # Show table in message box
            QMessageBox.information(self, "Test Successful", 
                                   f"Successfully read data from sheet '{inventory_sheet}'.\n"
                                   f"First 5 rows shown below.\n\n"
                                   f"Selected columns validated successfully!")
            
            # Show table in separate window
            table.show()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error testing configuration:\n{str(e)}")
            
    def save_configuration(self):
        """Save configuration to config.json"""
        try:
            # Get selected sheets
            inventory_sheet = None
            for button in self.inventory_sheet_group.buttons():
                if button.isChecked():
                    inventory_sheet = button.text()
                    break
            if not inventory_sheet and self.custom_inventory_sheet.text():
                inventory_sheet = self.custom_inventory_sheet.text()
                
            other_sheet = None
            for button in self.other_sheet_group.buttons():
                if button.isChecked():
                    other_sheet = button.text()
                    break
            if not other_sheet and self.custom_other_sheet.text():
                other_sheet = self.custom_other_sheet.text()
                
            if not inventory_sheet or not other_sheet:
                QMessageBox.critical(self, "Error", "Please select both inventory and other sheets!")
                return
                
            # Build configuration
            config = {
                "excel": {
                    "filePath": self.file_path_edit.text(),
                    "sheets": {
                        "inventory": inventory_sheet,
                        "other": other_sheet
                    },
                    "columns": {
                        "assetIdSearch": [
                            self.letter_to_index(self.asset_id1_combo.currentText()),
                            self.letter_to_index(self.asset_id2_combo.currentText()),
                            self.letter_to_index(self.asset_id3_combo.currentText())
                        ],
                        "assetName": self.letter_to_index(self.asset_name_combo.currentText()),
                        "assetDescription": self.letter_to_index(self.asset_desc_combo.currentText()),
                        "status": self.letter_to_index(self.status_combo.currentText()),
                        "location": self.letter_to_index(self.location_combo.currentText()),
                        "room": self.letter_to_index(self.room_combo.currentText()),
                        "markedCheck": self.letter_to_index(self.marked_check_combo.currentText())
                    },
                    "counting": {
                        "startRow": self.start_row_spin.value(),
                        "endRow": self.end_row_spin.value(),
                        "totalCount": self.end_row_spin.value() - self.start_row_spin.value() + 1
                    }
                },
                "server": {
                    "ngrokUrl": self.ngrok_url_edit.text()
                }
            }
            
            # Backup existing config.json
            config_path = os.path.join(os.path.dirname(__file__), "config.json")
            if os.path.exists(config_path):
                backup_name = f"config.json.backup.{datetime.now().strftime('%Y-%m-%d-%H%M%S')}"
                backup_path = os.path.join(os.path.dirname(__file__), backup_name)
                shutil.copy(config_path, backup_path)
                
            # Save config.json
            with open(config_path, 'w') as f:
                json.dump(config, f, indent=2)
                
            QMessageBox.information(self, "Success", 
                                   "Configuration saved successfully!\n\n"
                                   "The server URL will be automatically loaded from config.json.\n"
                                   "Please restart the app to apply changes.")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error saving configuration:\n{str(e)}")


def main():
    app = QApplication(sys.argv)
    window = ConfigWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
